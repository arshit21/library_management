import openpyxl

wb = openpyxl.load_workbook('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')
ws = wb.active
max_row = ws.max_row

user_type = int(input('Enter 0 if you are a Librarian or 1 if you are a Basic User: \n'))
if user_type == 0:
    pass
elif user_type == 1:
    pass
else:
    print('User Type not defined')
    quit()


class User():
    def __init__(self, user_type):
        if user_type == 0:
            self.user_type = 'Librarian'
        elif user_type == 1:
            self.user_type = 'Basic User'

user = User(user_type)

class Book():
    def __init__(self, id):
        self.author = ws.cell(row = id + 1, column = 3).value
        self.name = ws.cell(row = id + 1, column = 2).value
        self.ISBN = ws.cell(row = id + 1, column = 4).value
        self.issued_status = ws.cell(row = id + 1, column = 5).value
        self.id = id
        self.shelf = ws.cell(row = id + 1, column = 6).value

    def details(self):
        print(f'''
        name: {self.name}
        author: {self.author}
        ISBN: {self.ISBN}
        Is it Issued: {self.issued_status}
        ''')

    def edit_details(self):
        ques = input('Do you want to change the book name(Answer Yes or No): \n')
        if ques == 'Yes':
            ws.cell(row = self.id + 1, column = 2).value = input('Enter New name\n')
            self.name = ws.cell(row = self.id + 1, column = 2).value
            wb.save('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')
        else:
            pass
        ques_2 = input('Do you want to change the book author(Answer Yes or No): \n')
        if ques_2 == 'Yes':
            ws.cell(row = self.id + 1, column = 3).value = input('Enter New Author name\n')
            self.author = ws.cell(row = self.id + 1, column = 3).value
            wb.save('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')
        else:
            pass
        ques_3 = input("Do you want to change the book's ISBN(Answer Yes or No): \n")
        if ques_3 == 'Yes':
            ws.cell(row = self.id + 1, column = 4).value = input('Enter New ISBN \n')
            self.ISBN = ws.cell(row = self.id + 1, column = 4).value
            wb.save('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')
        else:
            pass
        print('The changes have been saved')

    def borrow_book(self):
        if self.issued_status == 'YES':
            print(f'The book {self.name} is already issued, cannot issue it now')
        elif self.issued_status == 'RESERVED':
            print(f'The book {self.name} is reserved, cannot issue it now')
        else:
            print(f'The book {self.name} is issued to you')
            ws.cell(row = self.id + 1, column = 5).value = 'YES'
            wb.save('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')

    def return_book(self):
        if self.issued_status == 'NO':
            print(f'The book {self.name} is not issued/reserved to/for anyone')
        else:
            print(f'The book {self.name} is returned by you')
            ws.cell(row = self.id + 1, column = 5).value = 'NO'
            wb.save('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')

    def reserve_book(self):
        if self.issued_status == 'YES':
            print(f'The book {self.name} is already issued, cannot reserve it now')
        elif self.issued_status == 'RESERVED':
            print(f'The book {self.name} is already reserved, cannot reserve it now')
        else:
            print(f'The book {self.name} is reserved for you')
            ws.cell(row = self.id + 1, column = 5).value = 'RESERVED'
            wb.save('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')

class Shelf():
    def __init__(self, s_number, User):
        self.number = s_number
        self.user_type = User.user_type

    def add_book(self, Book):
        if self.user_type == 'Basic User':
            print('Only Librarians can add a book to a shelf')
        else:
            ele = ws.cell(row = Book.id + 1, column = 6)
            if int(ele.value) == 0:
                print(f'The book {Book.name} has been added to shelf number {self.number}')
                ele.value = self.number
                wb.save('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')
            else:
                print(f'The book {Book.name} is already in shelf number {ele.value}')

    def rem_book(self, Book):
        if self.user_type == 'Basic User':
            print('Only Librarians can remove a book form a shelf')
        else:  
            ele = ws.cell(row = Book.id + 1, column = 6)
            if int(ele.value) == 0:
                print(f'The book {Book.name} is not in any shelf')
            elif int(ele.value) != self.number:
                print(f'The book {Book.name} is not in this shelf, it is in shelf number {Book.shelf}')
            else:
                print(f'The book {Book.name} has been removed from this shelf')
                ele.value = 0
                wb.save('/Users/arshitchaudhary/Desktop/python/Oop/dvm_task_1/books.xlsx')

    def show_catalog(self):
        print(f'In shelf number {self.number}, the following books are there:')
        for i in range(2, max_row + 1):
            ele = ws.cell(row = i, column = 6)
            if ele.value == self.number:
                print(f'{ws.cell(row = i, column = 2).value} written by {ws.cell(row = i, column = 3).value}')
        

    def get_books_count(self):
        count = 0
        for i in range(2, max_row + 1):
            ele = ws.cell(row = i, column = 6)
            if int(ele.value) == self.number:
                count += 1
        print(f'The total number of books in shelf number {self.number} are {count}')

        
id_no = int(input('Enter the id of the desired book: \n'))
book = Book(id_no)
shelf_no = int(input('Enter the desired shelf number: \n'))
shelf = Shelf(shelf_no, user)
print(f'You are a {user.user_type}')
if user.user_type == 'Librarian':
    method = int(input(f'''
    The following methods are available to you:
    enter 1 if you want to add the current book to the shelf
    enter 2 if you want to remove the current book from the shelf
    enter 3 if you want to see the catalog of books in the current shelf
    enter 4 if you want to see the total number of books in the current shelf 
    enter 5 if you want to change the book details
    enter 6 if you want to view the book details \n'''))

    if method == 1:
        shelf.add_book(book)
    elif method == 2:
        shelf.rem_book(book)
    elif method == 3:
        shelf.show_catalog()
    elif method == 4:
        shelf.get_books_count()
    elif method == 5:
        book.edit_details()
    elif method == 6:
        book.details()
    else:
        print('Invalid input')
        quit()

        
else:
    method = int(input(f'''
    enter 1 if you want to see the catalog of books in the current shelf
    enter 2 if you want to see the total number of books in the current shelf
    enter 3 if you want to reserve the book
    enter 4 if you want to return the book
    enter 5 if you want to issue the book
    enter 6 if you want to see the book details\n'''))

    if method == 1:
        shelf.show_catalog()
    elif method == 2:
        shelf.get_books_count()
    elif method == 3:
        book.reserve_book()
    elif method == 4:
        book.return_book()
    elif method == 5:
        book.borrow_book()
    elif method == 6:
        book.details()
    else:
        print('Invalid input')
        quit()